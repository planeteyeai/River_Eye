"""Convert flood_water_timeseries.xlsx into compact period JSON for the map heatmap."""

from collections import defaultdict
from pathlib import Path
import json
import sys

from openpyxl import load_workbook

src = Path(
    sys.argv[1]
    if len(sys.argv) > 1
    else r"c:\Users\Kunal.Desale\Downloads\d71568218a3e490db80433414967a4e5.xlsx"
)
asset = Path(__file__).resolve().parents[1] / "public" / "asset"

wb = load_workbook(src, read_only=True, data_only=True)

periods = []
ws = wb["areas"]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    periods.append(
        {
            "id": i - 1,
            "pre_date": row[0],
            "post_date": row[1],
            "water_area_ha": round(float(row[2]), 2),
            "flood_area_ha": round(float(row[3]), 2),
        }
    )

key_to_id = {(p["pre_date"], p["post_date"]): p["id"] for p in periods}
grouped = {p["id"]: {"flood": [], "water": []} for p in periods}

ws = wb["lat_lon"]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    pre, post, lat, lon, cls = row
    pid = key_to_id[(pre, post)]
    grouped[pid][cls].append([round(float(lon), 6), round(float(lat), 6)])

wb.close()

lons, lats = [], []
for pid, classes in grouped.items():
    n_flood = len(classes["flood"])
    n_water = len(classes["water"])
    periods[pid]["n_flood"] = n_flood
    periods[pid]["n_water"] = n_water
    periods[pid]["points"] = f"/asset/mula-mutha-flood-water-{pid}.json"
    for pair in classes["flood"] + classes["water"]:
        lons.append(pair[0])
        lats.append(pair[1])
    out = asset / f"mula-mutha-flood-water-{pid}.json"
    out.write_text(json.dumps(classes, separators=(",", ":")), encoding="utf-8")
    print("wrote", out.name, out.stat().st_size, "flood", n_flood, "water", n_water)

peak = max(periods, key=lambda p: p["flood_area_ha"])
index = {
    "name": "Mula-Mutha flood water timeseries",
    "source_file": "flood_water_timeseries.xlsx",
    "source_id": "d71568218a3e490db80433414967a4e5.xlsx",
    "captured": f"{periods[0]['pre_date']} to {periods[-1]['post_date']}",
    "kind": "Estimated",
    "note": (
        "Classed water and flood sample points for seven pre/post image pairs. "
        "Heatmap is point density, not a surveyed flood outline. Areas (ha) come "
        "from the workbook's areas sheet."
    ),
    "bounds": {
        "west": min(lons),
        "south": min(lats),
        "east": max(lons),
        "north": max(lats),
    },
    "default_period": peak["id"],
    "classes": {
        "water": {"label": "Surface water", "color": "#2f9bd6"},
        "flood": {"label": "Flood water", "color": "#c2372a"},
    },
    "periods": periods,
}

index_path = asset / "mula-mutha-flood-water.json"
index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")
print("wrote", index_path.name, "periods", len(periods), "default", peak["id"])
